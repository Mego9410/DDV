import argparse
import json
import os
import re
import time


LABELS: dict[str, list[str]] = {
    "practice_name": [r"practice\s*name", r"client\s*name", r"practice"],
    "effective_date": [
        r"effective\s*date",
        r"valuation\s*date",
        r"calc\s*eff",
        r"\bas\s*at\b",
        r"\bdate\b",
    ],
    "gross_fees": [r"gross\s*fees", r"total\s*fees", r"turnover"],
    "nhs_fees": [r"nhs\s*fees", r"\bnhs\b"],
    "private_fees": [r"private\s*fees", r"\bprivate\b"],
    "adjusted_profit": [
        r"adjusted\s*profit",
        r"maintainable\s*profit",
        r"normalized\s*profit",
        r"net\s*profit",
        r"recon(?:ciled)?\s*profit",
        r"\brnp\b",
        r"\bsurplus\b",
    ],
    "ebitda": [r"\bebitda\b", r"earnings\s*before"],
    "owner_remuneration": [
        r"principal\s*salary",
        r"owner\s*salary",
        r"remuneration",
        r"notional\s*salary",
    ],
    "rent": [r"\brent\b"],
    "rent_review": [r"rent\s*review", r"\brpi\b", r"\bcpi\b", r"index"],
    "goodwill": [r"\bgoodwill\b", r"\bgw\b", r"practice\s*value\s*of\s*gw"],
    "valuation": [
        r"\bvaluation\b",
        r"enterprise\s*value",
        r"total\s*value",
        r"practice\s*value",
        r"\bpv\b",
    ],
    "multiple": [r"\bmultiple\b", r"\bmult\b", r"\bmultiplier\b", r"\btimes\b", r"\bx\s*[0-9]"],
}


def _score_match(text: str | None, key: str) -> bool:
    t = (text or "").strip().lower()
    if not t:
        return False
    return any(re.search(pat, t) for pat in LABELS[key])


_DATE_RE = re.compile(
    r"(?P<d>0?[1-9]|[12][0-9]|3[01])(?P<sep>[./-])(?P<m>0?[1-9]|1[0-2])(?P=sep)(?P<y>\d{2}|\d{4})"
)


def _extract_date_from_text(text: str) -> str | None:
    m = _DATE_RE.search(text)
    if not m:
        return None
    d = int(m.group("d"))
    mo = int(m.group("m"))
    y = m.group("y")
    if len(y) == 2:
        y = "20" + y
    return f"{int(y):04d}-{mo:02d}-{d:02d}"


def _is_number(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _profile_value_neighbors_xlsx(ws, r: int, c: int, *, look_right: int = 12):
    """Yield (cell_ref, value) from cells to the right."""
    import openpyxl

    for dc in range(1, look_right + 1):
        v = ws.cell(r, c + dc).value
        if _is_number(v):
            col = openpyxl.utils.get_column_letter(c + dc)
            yield (f"{col}{r}", float(v))


def _profile_value_neighbors_xls(sh, r: int, c: int, *, look_right: int = 12):
    for dc in range(1, look_right + 1):
        if c + dc >= sh.ncols:
            break
        v = sh.cell_value(r, c + dc)
        if _is_number(v):
            yield (f"R{r+1}C{c+dc+1}", float(v))


def _summarize_matches(matches: dict[str, list[dict]]) -> dict[str, dict]:
    summary: dict[str, dict] = {}
    for k, arr in matches.items():
        if not arr:
            summary[k] = {"found": 0, "confidence": 0.0}
            continue
        sheets = len({x["sheet"] for x in arr})
        cells = len(arr)
        conf = min(1.0, 0.25 + 0.15 * sheets + 0.05 * min(cells, 10))
        summary[k] = {"found": cells, "confidence": round(conf, 2)}
    return summary


def _profile_xlsx(path: str, *, max_row: int, max_col: int) -> dict:
    import openpyxl

    # Use cached calculated values when present (critical for “multiple” etc),
    # while still scanning label strings normally.
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    info = {
        "type": "xlsx",
        "sheets": [],
        "matches": {k: [] for k in LABELS},
        "value_candidates": {k: [] for k in LABELS},
        "derived_candidates": {"effective_date": [], "multiple": []},
    }
    for ws in wb.worksheets:
        info["sheets"].append(
            {"name": ws.title, "max_row": ws.max_row or 0, "max_col": ws.max_column or 0}
        )

        # Dates often live in sheet names (eg CALCEFF_16.11.2020)
        sd = _extract_date_from_text(ws.title)
        if sd:
            info["derived_candidates"]["effective_date"].append(
                {"source": "sheet_name", "sheet": ws.title, "date": sd}
            )

        rmax = min(ws.max_row or 1, max_row)
        cmax = min(ws.max_column or 1, max_col)
        for r in range(1, rmax + 1):
            for c in range(1, cmax + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str):
                    for k in LABELS:
                        if _score_match(v, k):
                            col = openpyxl.utils.get_column_letter(c)
                            info["matches"][k].append(
                                {"sheet": ws.title, "cell": f"{col}{r}", "text": v[:120]}
                            )
                            # If label-like cell is found, try to pick up a numeric/date value to the right
                            for v_cell, v_num in _profile_value_neighbors_xlsx(ws, r, c):
                                info["value_candidates"][k].append(
                                    {
                                        "sheet": ws.title,
                                        "label_cell": f"{col}{r}",
                                        "value_cell": v_cell,
                                        "value": v_num,
                                        "label": v[:120],
                                    }
                                )

                            # Heuristic: the valuation methods table often has a numeric multiple/value
                            # in the cell(s) immediately to the right, but without the word "multiple".
                            lv = v.lower()
                            if any(
                                s in lv
                                for s in [
                                    "gw only",
                                    "gross fees",
                                    "recon profit",
                                    "reconciled profit",
                                    "based on last 12 m",
                                    "based on average of last 3",
                                ]
                            ):
                                for v_cell, v_num in _profile_value_neighbors_xlsx(ws, r, c):
                                    if 0.5 <= v_num <= 3.0:
                                        info["derived_candidates"]["multiple"].append(
                                            {
                                                "source": "valuation_methods_row",
                                                "sheet": ws.title,
                                                "label_cell": f"{col}{r}",
                                                "value_cell": v_cell,
                                                "value": v_num,
                                                "label": v[:120],
                                            }
                                        )
    return info


def _profile_xls(path: str, *, max_row: int, max_col: int) -> dict:
    import xlrd

    book = xlrd.open_workbook(path)
    info = {
        "type": "xls",
        "sheets": [],
        "matches": {k: [] for k in LABELS},
        "value_candidates": {k: [] for k in LABELS},
        "derived_candidates": {"effective_date": [], "multiple": []},
    }
    for sh in book.sheets():
        info["sheets"].append({"name": sh.name, "rows": sh.nrows, "cols": sh.ncols})

        sd = _extract_date_from_text(sh.name)
        if sd:
            info["derived_candidates"]["effective_date"].append(
                {"source": "sheet_name", "sheet": sh.name, "date": sd}
            )

        rmax = min(sh.nrows, max_row)
        cmax = min(sh.ncols, max_col)
        for r in range(rmax):
            for c in range(cmax):
                v = sh.cell_value(r, c)
                if isinstance(v, str):
                    for k in LABELS:
                        if _score_match(v, k):
                            info["matches"][k].append(
                                {"sheet": sh.name, "cell": f"R{r+1}C{c+1}", "text": v[:120]}
                            )
                            for v_cell, v_num in _profile_value_neighbors_xls(sh, r, c):
                                info["value_candidates"][k].append(
                                    {
                                        "sheet": sh.name,
                                        "label_cell": f"R{r+1}C{c+1}",
                                        "value_cell": v_cell,
                                        "value": v_num,
                                        "label": v[:120],
                                    }
                                )

                            lv = v.lower()
                            if any(
                                s in lv
                                for s in [
                                    "gw only",
                                    "gross fees",
                                    "recon profit",
                                    "reconciled profit",
                                    "based on last 12 m",
                                    "based on average of last 3",
                                ]
                            ):
                                for v_cell, v_num in _profile_value_neighbors_xls(sh, r, c):
                                    if 0.5 <= v_num <= 3.0:
                                        info["derived_candidates"]["multiple"].append(
                                            {
                                                "source": "valuation_methods_row",
                                                "sheet": sh.name,
                                                "label_cell": f"R{r+1}C{c+1}",
                                                "value_cell": v_cell,
                                                "value": v_num,
                                                "label": v[:120],
                                            }
                                        )
    return info


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--paths-json", required=True, help="JSON array of absolute file paths.")
    parser.add_argument("--max-row", type=int, default=200)
    parser.add_argument("--max-col", type=int, default=30)
    parser.add_argument("--out", required=False, help="Write JSON results to this file.")
    args = parser.parse_args()

    with open(args.paths_json, "r", encoding="utf-8") as f:
        paths = json.load(f)
    if not isinstance(paths, list) or not all(isinstance(p, str) for p in paths):
        raise SystemExit("--paths-json must contain a JSON array of strings")

    results: list[dict] = []
    for idx, p in enumerate(paths, start=1):
        t0 = time.time()
        base = os.path.basename(p)
        print(f"[{idx}/{len(paths)}] {base}", flush=True)

        rec: dict = {"path": p, "exists": os.path.exists(p)}
        if not rec["exists"]:
            rec["error"] = "not_found"
            results.append(rec)
            continue

        try:
            # Dates also frequently live in file names (eg CALCEFF_16.11.2020 / 06-10-21)
            fn_date = _extract_date_from_text(os.path.basename(p))
            if p.lower().endswith(".xlsx"):
                info = _profile_xlsx(p, max_row=args.max_row, max_col=args.max_col)
            elif p.lower().endswith(".xls"):
                info = _profile_xls(p, max_row=args.max_row, max_col=args.max_col)
            else:
                rec["error"] = "unsupported_ext"
                results.append(rec)
                continue
            rec["type"] = info["type"]
            rec["sheets"] = info["sheets"]
            # Upgrade confidence if we can see a value next to a label, or a derived date/multiple exists.
            ms = _summarize_matches(info["matches"])
            for k, arr in info.get("value_candidates", {}).items():
                if arr:
                    ms[k]["confidence"] = max(ms[k]["confidence"], 0.85)

            if info.get("value_candidates", {}).get("effective_date"):
                ms["effective_date"]["confidence"] = max(ms["effective_date"]["confidence"], 0.8)
            if fn_date or info.get("derived_candidates", {}).get("effective_date"):
                ms["effective_date"]["confidence"] = max(ms["effective_date"]["confidence"], 0.7)

            if info.get("value_candidates", {}).get("multiple") or info.get("derived_candidates", {}).get(
                "multiple"
            ):
                ms["multiple"]["confidence"] = max(ms["multiple"]["confidence"], 0.85)

            rec["match_summary"] = ms
            rec["sample_matches"] = {k: info["matches"][k][:2] for k in info["matches"]}
            rec["derived_candidates"] = info.get("derived_candidates", {})
            rec["value_candidates_sample"] = {
                k: info.get("value_candidates", {}).get(k, [])[:2] for k in LABELS
            }
        except Exception as e:
            rec["error"] = f"parse_failed: {type(e).__name__}: {e}"

        rec["elapsed_s"] = round(time.time() - t0, 2)
        results.append(rec)

    out_json = json.dumps(results, indent=2)
    if args.out:
        with open(args.out, "w", encoding="utf-8") as f:
            f.write(out_json)
    else:
        print("\n===JSON===\n")
        print(out_json)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

